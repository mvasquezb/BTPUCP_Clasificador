import csv
import openpyxl
from textprocessor import TextProcessor
from pathlib import Path as OSPath
from utils import open_filename
from collections import OrderedDict


class DataLoader:
    def __init__(self, root='.'):
        self.root = root
        self.textProcessor = TextProcessor()

    def clean_dataset(self, dataset, inplace=False):
        """
        Merge offer attributes and preprocess text
        """
        dataset = OrderedDict(
            sorted(dict.items(), key=lambda t: t[0])
        ) if inplace else dataset
        new_data = dataset if inplace else OrderedDict(dataset.copy())
        for key, offer in new_data.items():
            new_data[key] = ' '.join(
                set(self.textProcessor.tokenize(' '.join(offer)))
            )

        return new_data

    def get_labelled_data(self, carrera):
        "Lee de un archivo la etiqueta que le corresponde a cada oferta."

        labelledData = {}
        Id = 0
        categoria = 1

        with open(str(OSPath(
            self.root, carrera, 'dataEtiquetadaABCD.txt')
        )) as f:
            offers = csv.reader(f)
            for offer in offers:
                labelledData[int(offer[Id])] = offer[categoria].lower()
        return labelledData

    def get_dataset(self, labelledData, carrera):
        """
        Lee las ofertas desde el archivo ofertasCSV.txt
        y las inserta en una lista.
        """

        Id = 0
        dataset = {}
        with open(str(OSPath(
            self.root, carrera, 'dataEntrenamiento.csv')
        )) as f:
            offers = csv.reader(f)
            # Skip header
            next(offers)
            for offer in offers:
                if int(offer[Id]) in labelledData.keys():
                    dataset[int(offer[Id])] = offer[1:]

        return self.clean_dataset(dataset)

    def get_targets(self, career, encoding='UTF-8'):
        "Obtiene las categorías que pertenecen a la carrera escogida."

        targets = []

        with open(str(
            OSPath(self.root, career, 'categorias.txt')
        ), encoding=encoding) as f:
            for target in f:
                target = target.lower().strip()
                targets.append(target.replace("\n", ""))
        targets.sort()
        return targets

    def get_features(self, career, encoding='ISO-8859-1'):
        offer_dict = {}

        with open(str(
            OSPath(self.root, career, "diccProfeABCD/diccionarios.txt")
        ), encoding=encoding) as f1:

            for target in f1:
                target = target.replace("\n", "")
                print("Diccionarios-", target)
                with open(str(
                    OSPath(self.root, career, "diccProfeABCD/", target)
                ), encoding=encoding) as f2:

                    target = (target[:-len('.txt')]
                              if target.endswith('.txt') else target)

                    if target not in offer_dict.keys():
                        offer_dict[target] = set()

                    for feature in f2:
                        feature = self.textProcessor.preprocess(feature)
                        feature = self.textProcessor.stem(feature)
                        offer_dict[target].add(feature)

        for cat in offer_dict.keys():
            offer_dict[cat] = list(offer_dict[cat])

        return offer_dict

    def get_data_for_career(self, career):
        labelled_data = self.get_labelled_data(career)
        dataset = self.get_dataset(labelled_data, career)
        targets = self.get_targets(career)
        features = self.get_features(career)
        return {
            'labelled_data': labelled_data,
            'dataset': dataset,
            'features': features,
            'target_names': targets,
        }

    def printPredicted(self,
                       labelledData,
                       predicted,
                       filename,
                       carrera):
        division_Carpetas = filename.split("/")
        filename = division_Carpetas[len(division_Carpetas) - 1]
        first = 0
        file_base = filename.split(".")[first]
        indID = 0
        with open(str(
            OSPath(self.root, carrera, "DataClasificada_" + file_base + ".txt")
        ), 'w') as f:
            for i in range(len(labelledData)):
                f.write("%s: %s\n" % (labelledData[i][indID], predicted[i]))

    def print_dictionaries(self,
                           cat_word_count,
                           categorias,
                           filename,
                           carrera):
        division_Carpetas = filename.split("/")
        filename = division_Carpetas[len(division_Carpetas) - 1]
        firstElement = 0
        file_base = filename.split(".")[firstElement]
        for categoriaEtiqueta in categorias:
            for categoria in categorias:
                with open(str(
                    OSPath(self.root,
                           carrera, "Diccionario_" + file_base + "_" +
                           categoriaEtiqueta + "_" + categoria + ".txt")
                ), 'w') as f:
                    for palabra in sorted(cat_word_count[categoriaEtiqueta][categoria].keys()):
                        f.write("%s: %d\n" % (
                            palabra, cat_word_count[categoriaEtiqueta][categoria][palabra]))

    def _read_excel_dataset(self, filename):
        dataset = {}

        wb = openpyxl.load_workbook(filename)
        sheets = wb.get_sheet_names()
        offer_sheet = wb.get_sheet_by_name(sheets[0])
        max_rows = offer_sheet.max_row + 1
        max_columns = offer_sheet.max_column + 1

        for offer_number in range(2, max_rows):
            offer_id = int(offer_sheet.cell(row=offer_number, column=1).value)
            offer = [
                str(offer_sheet.cell(
                    row=offer_number,
                    column=col_number
                ).value)

                for col_number in range(2, max_columns)
            ]
            dataset[offer_id] = offer

        return dataset

    def _read_csv_dataset(self, filename):
        dataset = []

        return dataset

    def get_unlabelled_data(self, filename):
        "Se lee un archivo Excel con las ofertas a clasificar"

        with open_filename(str(OSPath(self.root, filename)), 'rb') as f:
            path = OSPath(f.name)
            if path.suffix == '.xlsx':
                dataset = self._read_excel_dataset(f)
            else:
                # Assume csv
                dataset = self._read_csv_dataset(f)

        return self.clean_dataset(dataset)
