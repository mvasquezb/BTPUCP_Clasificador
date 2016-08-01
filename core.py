from nltk.corpus import stopwords
import csv
import operator
import codecs
import string
import math
import unicodedata
import treetaggerwrapper
import openpyxl
import re
import numpy
import warnings
from nltk.stem.snowball import SpanishStemmer

import matplotlib.pyplot as plt
from sklearn import datasets, svm, metrics
from sklearn.linear_model import LogisticRegression
from sklearn.cross_validation import train_test_split
from sklearn import cross_validation

from clasificador_categorias import Maximizador

class TextProcessor:
    lemmatizer=None
    stopEnglish=None
    stopSpanish=None
    spanishStemmer=None

    def __init__(self):
        self.lemmatizer = treetaggerwrapper.TreeTagger(TAGLANG='es')
        self.stopEnglish = stopwords.words('english')
        self.stopSpanish = stopwords.words('spanish')
        self.stopSpanish.append('y/o')
        self.spanishStemmer=SpanishStemmer()

    def _remove_numbers(self, text):
        "Elimina los números del texto"

        return ''.join([letter for letter in text if not letter.isdigit()])

    def _remove_punctuation(self, text):
        "Elimina los signos de puntuacion del texto"

        regex = re.compile('[%s]' % re.escape(string.punctuation))
        return regex.sub(' ', text)

    def preprocessText(self,text):
        text=text.lower()
        text=self._remove_punctuation(text)
        text=self._remove_numbers(text)
        return text

    def lematizeText(self,text):
        newText = ""
        firstElement = 0
        firstWord=True
        for word in text.split():
            if word not in self.stopEnglish and word not in self.stopSpanish:
                word = word.replace("\ufeff", "")
                lemmaResult = self.lemmatizer.tag_text(word)
                # Return [[word,type of word, lemma]]
                if (len(lemmaResult) != 0):
                    word = lemmaResult[firstElement].split()[2]
                    if firstWord:
                        newText += word
                        firstWord = False
                    else:
                        newText += " " + word
        return newText

    def stemText(self,text):
        newText = ""
        firstWord = True
        for word in text.split():
            if word not in self.stopEnglish and word not in self.stopSpanish:
                word = word.replace("\ufeff", "")
                wordStemmed = self.spanishStemmer.stem(word)
                if firstWord:
                    newText += wordStemmed
                    firstWord = False
                else:
                    newText += " " + wordStemmed
        return newText


class Input_Output:

    procesadorTextos=None

    def __init__(self):
        self.procesadorTextos=TextProcessor()

    def limpiarDataset(self, dataset):
        """Quita los signos de puntuación, pasa todo a minúsculas,
        quita los números, elimina los stopwords(inglés y español)
        y pasa todas las palabras a forma raíz."""

        Id = 0
        for oferta in dataset:
            oferta[Id] = oferta[Id].replace("\ufeff", "")
            numAtributos = len(oferta)
            for atributo in range(1, numAtributos):
                oferta[atributo] = self.procesadorTextos.preprocessText(oferta[atributo])
                #oferta[atributo] = self.procesadorTextos.lematizeText(oferta[atributo])
                oferta[atributo] = self.procesadorTextos.stemText(oferta[atributo])

        return dataset

    def obtenerCategorias(self, carrera):
        "Obtiene las categorías que pertenecen a la carrera escogida."

        categorias = []

        with open(carrera + '/categorias.txt') as f:
            lineas = f.readlines()
            for categoria in lineas:
                categoria = categoria.lower()
                categorias.append(categoria.replace("\n", ""))
        categorias.sort()
        return categorias

    def obtenerDataEtiquetada(self,carrera):
        "Lee de un archivo la etiqueta que le corresponde a cada oferta."

        dataEtiquetada={}
        Id=0
        categoria=1
        with open(carrera+'/dataEtiquetadaABCD.txt','r') as f:
            ofertas=csv.reader(f)
            for oferta in ofertas:
                dataEtiquetada[int(oferta[Id])]=oferta[categoria].lower()
        return dataEtiquetada

    def obtenerDataset(self,dataEtiquetada,carrera):
        "Lee las ofertas desde el archivo ofertasCSV.txt y las inserta en una lista."

        Id=0
        dataset=[]
        with open(carrera+'/dataEntrenamiento.txt','r') as f:
            ofertas = csv.reader(f)
            for oferta in ofertas:
                if int(oferta[Id]) in dataEtiquetada.keys():
                    dataset.append(oferta)
        return self.limpiarDataset(dataset)

    def obtenerDiccionario(self,carrera):
        dicc_ofertas = dict()

        with open(carrera + "/diccProfeABCD/diccionarios.txt", 'r', -1, 'utf-8') as f1:
            for categoria in f1.read().splitlines():
                categoria = categoria.replace("\ufeff", "")
                with open(carrera + "/diccProfeABCD/" + categoria, 'r') as f2:
                    categoria = categoria.replace(".txt", "")
                    if categoria not in dicc_ofertas.keys():
                        dicc_ofertas[categoria] = set()

                    for linea in f2.read().splitlines():
                        linea = self.procesadorTextos.preprocessText(linea)
                        #cadena_procesada = self.procesadorTextos.lematizeText(linea)
                        cadena_procesada = self.procesadorTextos.stemText(linea)
                        dicc_ofertas[categoria].add(cadena_procesada)

        for cat in dicc_ofertas.keys():
            dicc_ofertas[cat] = list(dicc_ofertas[cat])

        return dicc_ofertas

    def obtenerUtilidades(self,carrera):
        dataEtiquetada = self.obtenerDataEtiquetada(carrera)
        dataset = self.obtenerDataset(dataEtiquetada, carrera)
        diccionario = self.obtenerDiccionario(carrera)
        categorias = self.obtenerCategorias(carrera)
        return dataEtiquetada,dataset,diccionario,categorias

    def _imprimirPredicted(self,datasetClasificado,predicted,filename,carrera):
        division_Carpetas=filename.split("/")
        nombArchivo_Extension=division_Carpetas[len(division_Carpetas)-1]
        firstElement=0
        nombArchivo=nombArchivo_Extension.split(".")[firstElement]
        indID=0
        with open(carrera+"/DataClasificada_"+nombArchivo+".txt",'w') as f:
            for i in range(len(datasetClasificado)):
                f.write("%s: %s\n"%(datasetClasificado[i][indID],predicted[i]))

    def _imprimirDiccionarios(self,conteo_Categorias_Palabras,categorias,filename,carrera):
        division_Carpetas = filename.split("/")
        nombArchivo_Extension = division_Carpetas[len(division_Carpetas) - 1]
        firstElement = 0
        nombArchivo = nombArchivo_Extension.split(".")[firstElement]
        for categoriaEtiqueta in categorias:
            for categoria in categorias:
                with open(carrera + "/Diccionario_" + nombArchivo + "_"+categoriaEtiqueta+"_"+categoria+".txt", 'w') as f:
                    for palabra in sorted(conteo_Categorias_Palabras[categoriaEtiqueta][categoria].keys()):
                        f.write("%s: %d\n"%(palabra,conteo_Categorias_Palabras[categoriaEtiqueta][categoria][palabra]))

    def obtenerDatasetAClasificar(self, filename):
        "Se lee un archivo Excel con las ofertas a clasificar"

        dataset = []
        wb = openpyxl.load_workbook(filename)
        sheets = wb.get_sheet_names()
        sheetAviso = wb.get_sheet_by_name(sheets[0])
        maxFilas = sheetAviso.max_row + 1
        maxColumnas = sheetAviso.max_column + 1
        for num_oferta in range(2, maxFilas):
            fila = []
            for nColumna in range(1, maxColumnas):
                fila.append(str(sheetAviso.cell(row=num_oferta, column=nColumna).value))
            dataset.append(fila)

        return self.limpiarDataset(dataset)


class Core():
    carrera=""
    categorias=list()
    input_output=None

    def __init__(self,carrera):
        self.carrera=carrera
        self.input_output=Input_Output()

    def _count_word(self,text,search):
        result=re.findall('\\b'+search+'\\b',text,flags=re.IGNORECASE)
        return len(result)

    def fill_TF_IDF(self,dataset,diccionario,categorias):
        """Se llena un diccionario con 0's para luego llenarlos al momento de
        hacer el tf-idf"""
        
        TF_IDF={}

        Id=0
        for oferta in dataset:
            TF_IDF[int(oferta[Id])]={}  

        for ID in TF_IDF.keys():
            for cat in categorias:
                TF_IDF[ID][cat]=[]
                if(len(diccionario[cat])==0):
                    TF_IDF[ID][cat]=[0]
                    print("No hay palabras de la categoria " + cat)
                else:
                    TF_IDF[ID][cat]=[0]*len(diccionario[cat])                            

        return TF_IDF


    def calcular_idf(self,dataset,diccionario):
        """ Se calcula el idf de cada palabra por diccionario. Se sigue la fórmula idf = log(N/DF)
        N siendo la cantidad de avisos."""

        N=len(dataset)
        idf={}
        titulo=1
        descripcion=2
        requerimientos=3
        for cat in diccionario.keys():
            for palabra in diccionario[cat]:
                DF=int(0)
                for oferta in dataset:
                    if((palabra in oferta[titulo])or (palabra in oferta[descripcion]) or (palabra in oferta[requerimientos])):
                        DF+=1
                if(DF==0):
                    idf[palabra]=0
                else:
                    idf[palabra]=math.log(N/DF)

        return idf


    def normalizacion(self,dataset,diccionario,TF_IDF,idf):
        """Se normalizan los valores de TF_IDF dividiendo cada valor entre la
        palabra con la mayor frecuencia de la oferta."""
        
        max_frec={}
        indID=0
        titulo=1
        descripcion=2
        requerimientos=3
        for oferta in dataset:
            Id=int(oferta[indID])
            max_frec[Id]={}
            for cat in diccionario.keys():
                maximo=0 #se guarda el maximo, para hacer la normalizacion luego del calculo del tf-idf
                for palabra in diccionario[cat]:
                    tf = self._count_word(str(oferta[titulo]),str(palabra))
                    tf += self._count_word(str(oferta[descripcion]),str(palabra))
                    tf += self._count_word(str(oferta[requerimientos]),str(palabra))
                    if tf > maximo:
                        maximo=tf
                    TF_IDF[Id][cat][diccionario[cat].index(palabra)]=int(tf)*idf[palabra]
                        #se agrega la cantidad en la posición correspondiente
                max_frec[Id][cat]=maximo


        for Id in TF_IDF.keys():
            for cat in diccionario.keys():
                for indWord in range(len(TF_IDF[Id][cat])):
                    if max_frec[Id][cat]>0:
                        TF_IDF[Id][cat][indWord]/=max_frec[Id][cat]

        return TF_IDF


    def calcularTF_IDF(self,diccionario,dataset,categorias ):
        "Calcula el tf-idf para el dataset que se usará para las pruebas."
        #print(len(dataset))
        #comienza inicializacion de la estructura para calcular tf-idf
        TF_IDF=self.fill_TF_IDF(dataset,diccionario,categorias)

        idf=self.calcular_idf(dataset,diccionario)

        TF_IDF=self.normalizacion(dataset,diccionario,TF_IDF,idf)

        return TF_IDF

    def tranformarDataALista(self,TF_IDF,categorias):
        data=[]
        for Id in sorted(TF_IDF.keys()):
            oferta=[]
            for categoria in categorias:
                categoriasXoferta=[]
                for frec in TF_IDF[Id][categoria]:
                    categoriasXoferta.append(frec)
                oferta.append(categoriasXoferta)
            data.append(oferta)
        return data

    def obtenerCategoriasEsperadas(self,TF_IDF,dataEtiquetada):
        expected = []
        for Id in sorted(TF_IDF.keys()):
            expected.append(dataEtiquetada[Id])
        return expected

    def mostrarResultados(self,clasificador,expected,predicted):
        """Muestra la los resultados de precision, recall y f1-score para cada
           categoría y el promedio de estas. También muestra la matriz de confusión.
    """
        reporte="Classification report for classifier %s:\n%s\n"% (clasificador, metrics.classification_report(expected, predicted))
        matriz_confusion="Confusion matrix:\n%s" % metrics.confusion_matrix(expected, predicted)
        return reporte,matriz_confusion


    def entrenamiento(self,TF_IDF,dataEtiquetada,categorias):
        """Esta función entre a los clasificadores dentro del maximizador
           y prueba los datos usando cross validation."""

        data=self.tranformarDataALista(TF_IDF,categorias)
        expected=self.obtenerCategoriasEsperadas(TF_IDF,dataEtiquetada)
        
        clasificador=Maximizador(0,"defaultValue",categorias)

        #data contiene [[[],[],[],...,[]],[[],[],[],...,[]],...,[[],[],[],...,[]]]
        #Siendo el primer indice, la oferta, el segundo indice la categoria, y el tercer indice el TF_IDF
        predicted=cross_validation.cross_val_predict(clasificador,data,expected,cv=10)

        reporte,matriz_confusion=self.mostrarResultados(clasificador,expected,predicted)

        return clasificador,reporte,matriz_confusion        


    def crearMatriz(self,diccionario,categorias,predicted,ofertas):

        for categoria in categorias:
            print("Cantidad de",categoria,":",predicted.count(categoria))

        conteo_Categorias_Palabras={}
        for categoriaEtiqueta in categorias:
            conteo_Categorias_Palabras[categoriaEtiqueta]={}
            for categoria in categorias:
                conteo_Categorias_Palabras[categoriaEtiqueta][categoria] = {}

        matriz_conocimientos = [[0 for x in range(len(categorias))] for x in range(len(categorias))]

        titulo = 1
        descripcion = 2
        requerimientos = 3
        for numOferta in range(len(predicted)):
            categoriaEtiqueta=predicted[numOferta]
            indCategoriaEtiquetada=categorias.index(categoriaEtiqueta)

            for categoria in categorias:
                indCategoria=categorias.index(categoria)
                if categoriaEtiqueta != categoria:
                    alreadyCount=False
                else:
                    matriz_conocimientos[indCategoriaEtiquetada][indCategoria] += 1
                    alreadyCount = True

                for palabra in diccionario[categoria]:
                    if(self._count_word(ofertas[numOferta][titulo],palabra)>0 or self._count_word(ofertas[numOferta][descripcion],palabra)>0 or self._count_word(ofertas[numOferta][requerimientos],palabra)>0):
                        if palabra not in conteo_Categorias_Palabras[categoriaEtiqueta][categoria].keys():
                            conteo_Categorias_Palabras[categoriaEtiqueta][categoria][palabra] = 1
                        else:
                            conteo_Categorias_Palabras[categoriaEtiqueta][categoria][palabra] += 1
                        if(not(alreadyCount)):
                            matriz_conocimientos[indCategoriaEtiquetada][indCategoria]+=1
                            alreadyCount=True

        return matriz_conocimientos,conteo_Categorias_Palabras



    def predecir(self,clasificador, ofertas,diccionario,categorias):

        TF_IDF=self.calcularTF_IDF(diccionario,ofertas,categorias)

        data=self.tranformarDataALista(TF_IDF,categorias)
        predicted=clasificador.predict(data)

        matrizConocimientos,conteo_Categorias_Palabras=self.crearMatriz(diccionario,categorias,predicted,ofertas)

        return predicted,matrizConocimientos,conteo_Categorias_Palabras



    def EjecutarEntrenamiento(self):

        dataEtiquetada, dataset, diccionario, categorias=self.input_output.obtenerUtilidades(self.carrera)

        TF_IDF=self.calcularTF_IDF(diccionario,dataset,categorias)
        clasificador,res1,res2=self.entrenamiento(TF_IDF,dataEtiquetada,categorias)
        return res1,res2
        

    def EjecutarClasificacion(self,filename):
        dataEtiquetada, dataset, diccionario, categorias = self.input_output.obtenerUtilidades(self.carrera)

        TF_IDF=self.calcularTF_IDF(diccionario,dataset,categorias)
        clasificador,res1,res2=self.entrenamiento(TF_IDF,dataEtiquetada,categorias)
        print("Se finalizó la etapa de entrenamiento")
        datasetAClasificar=self.input_output.obtenerDatasetAClasificar(filename)
        predicted,mat_con,conteo_Categorias_Palabras=self.predecir(clasificador,datasetAClasificar,diccionario,categorias)
        self.input_output._imprimirPredicted(datasetAClasificar,predicted,filename,self.carrera)
        self.input_output._imprimirDiccionarios(conteo_Categorias_Palabras,categorias,filename,self.carrera)
        return mat_con

