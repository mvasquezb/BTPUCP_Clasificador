import csv
import openpyxl
from textprocessor import TextProcessor
from pathlib import Path as OSPath
from utils import open_filename


class DataLoader:

    procesadorTextos = None

    def __init__(self):
        self.procesadorTextos = TextProcessor()

    def clean_dataset(self, dataset):
        """Quita los signos de puntuación, pasa todo a minúsculas,
        quita los números, elimina los stopwords(inglés y español)
        y pasa todas las palabras a forma raíz."""

        Id = 0
        for oferta in dataset:
            oferta[Id] = oferta[Id].replace("\ufeff", "")
            numAtributos = len(oferta)
            for attr in range(1, numAtributos):
                oferta[attr] = self.procesadorTextos.preprocessText(
                    oferta[attr]
                )
                # oferta[attr] = self.procesadorTextos.lematizeText(oferta[attr])
                oferta[attr] = self.procesadorTextos.stemText(oferta[attr])

        return dataset

    def obtenerDataEtiquetada(self, carrera):
        "Lee de un archivo la etiqueta que le corresponde a cada oferta."

        dataEtiquetada = {}
        Id = 0
        categoria = 1

        with open(carrera + '/dataEtiquetadaABCD.txt') as f:
            ofertas = csv.reader(f)
            for oferta in ofertas:
                dataEtiquetada[int(oferta[Id])] = oferta[categoria].lower()
        return dataEtiquetada

    def obtenerDataset(self, dataEtiquetada, carrera):
        """
        Lee las ofertas desde el archivo ofertasCSV.txt
        y las inserta en una lista.
        """

        Id = 0
        dataset = []
        with open(carrera + '/dataEntrenamiento.csv') as f:
            ofertas = csv.reader(f)
            # Skip header
            next(ofertas)
            for oferta in ofertas:
                if int(oferta[Id]) in dataEtiquetada.keys():
                    dataset.append(oferta)
        return self.clean_dataset(dataset)

    def obtenerCategorias(self, carrera):
        "Obtiene las categorías que pertenecen a la carrera escogida."

        categorias = []

        with open(carrera + '/categorias.txt') as f:
            for categoria in f:
                categoria = categoria.lower()
                categorias.append(categoria.replace("\n", ""))
        categorias.sort()
        return categorias

    def get_dictionary(self, carrera):
        dicc_ofertas = dict()

        with open(carrera + "/diccProfeABCD/diccionarios.txt",
                  encoding='ISO-8859-1') as f1:
            lineas = f1.readlines()
            for categoria in lineas:
                categoria = categoria.replace("\n", "")
                print("Diccionarios-", categoria)
                with open(carrera + "/diccProfeABCD/" + categoria,
                          encoding='ISO-8859-1') as f2:
                    categoria = categoria.replace(".txt", "")
                    if categoria not in dicc_ofertas.keys():
                        dicc_ofertas[categoria] = set()

                    for linea in f2.read().splitlines():
                        linea = self.procesadorTextos.preprocessText(linea)
                        # cadena_procesada = self.procesadorTextos.lematizeText(linea)
                        cadena_procesada = self.procesadorTextos.stemText(
                            linea)
                        dicc_ofertas[categoria].add(cadena_procesada)

        for cat in dicc_ofertas.keys():
            dicc_ofertas[cat] = list(dicc_ofertas[cat])

        return dicc_ofertas

    def get_data_for_career(self, carrera):
        dataEtiquetada = self.obtenerDataEtiquetada(carrera)
        dataset = self.obtenerDataset(dataEtiquetada, carrera)
        categorias = self.obtenerCategorias(carrera)
        diccionario = self.get_dictionary(carrera)

        return dataEtiquetada, dataset, diccionario, categorias

    def printPredicted(self,
                       datasetClasificado,
                       predicted,
                       filename,
                       carrera):
        division_Carpetas = filename.split("/")
        nombArchivo_Extension = division_Carpetas[len(division_Carpetas) - 1]
        firstElement = 0
        nombArchivo = nombArchivo_Extension.split(".")[firstElement]
        indID = 0
        with open(carrera + "/DataClasificada_" +
                  nombArchivo + ".txt", 'w') as f:
            for i in range(len(datasetClasificado)):
                f.write("%s: %s\n" %
                        (datasetClasificado[i][indID], predicted[i]))

    def print_dictionaries(self,
                              cat_word_count,
                              categorias,
                              filename,
                              carrera):
        division_Carpetas = filename.split("/")
        nombArchivo_Extension = division_Carpetas[len(division_Carpetas) - 1]
        firstElement = 0
        nombArchivo = nombArchivo_Extension.split(".")[firstElement]
        for categoriaEtiqueta in categorias:
            for categoria in categorias:
                with open(carrera + "/Diccionario_" + nombArchivo + "_" +
                          categoriaEtiqueta + "_" +
                          categoria + ".txt", 'w') as f:
                    for palabra in sorted(cat_word_count[categoriaEtiqueta][categoria].keys()):
                        f.write("%s: %d\n" % (
                            palabra, cat_word_count[categoriaEtiqueta][categoria][palabra]))

    def _read_excel(self, filename):
        dataset = []

        wb = openpyxl.load_workbook(filename)
        sheets = wb.get_sheet_names()
        sheetAviso = wb.get_sheet_by_name(sheets[0])
        maxFilas = sheetAviso.max_row + 1
        maxColumnas = sheetAviso.max_column + 1

        for num_oferta in range(2, maxFilas):
            fila = []
            for nColumna in range(1, maxColumnas):
                fila.append(
                    str(sheetAviso.cell(row=num_oferta, column=nColumna).value)
                )
            dataset.append(fila)

        return dataset

    def _read_csv(self, filename):
        dataset = []

        return dataset

    def getUnlabelledData(self, filename):
        "Se lee un archivo Excel con las ofertas a clasificar"

        with open_filename(filename, 'rb') as f:
            path = OSPath(f.name)
            if path.suffix == '.xlsx':
                dataset = self._read_excel(f)
            else:
                # Assume csv
                dataset = self._read_csv(f)

        return self.clean_dataset(dataset)
